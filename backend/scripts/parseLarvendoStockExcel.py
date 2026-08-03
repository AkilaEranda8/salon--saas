#!/usr/bin/env python3
"""Parse Larvendo Stock.xlsx → products JSON for import."""
import json
import re
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Akila\Downloads\Copy of Larvendo Stock.xlsx")
OUT = Path(r"E:\salon_v1\backend\scripts\data\larvendo_stock_products.json")

CATEGORIES = {
    "colors", "developer", "shampo", "shampoo", "bleaching", "conditioner",
    "keratin", "mask", "cream", "gel", "pack", "serum", "toner", "cleanser",
    "scrub", "packet", "bean wax", "reducteur", "waving lotion", "smooth cream",
    "repair mask", "charcoal cream", "massage cream", "hydra liquid", "wax oil",
    "foot mask", "bubble packet", "twiser", "plastic form", "hair tretments",
    "hair treatments products", "kaeun treatment", "bond fusion 1", "hair mask",
    "gold hair mask", "keratin smooth oil", "keratin mask", "keratin shampoo",
    "keratin lotion", "facial korean brand", "foaming cleanser", "gentle cleanser",
    "purfying  toner", "soothing toner", "h2o massage cream", "protective emulsion",
    "face pack 1", "face pack 2", "face pack 3", "hyaluronic serum", "brightining serum",
    "facial scrub", "anti acne face pack", "face scrub", "glow pack", "gold facial pack",
    "melaless peel", "max hydra gel", "aciderm foam gel", "aciderm toner",
    "exo regentron toner", "exo regentron foamgel", "salieylicacid face waash",
    "blackhead white head remover", "liquorice mud pack", "face pack",
    "massage cream pedicure", "straigtiting cream & neautrilizer", "dreamron 30 volume",
}


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def unit_for(name, ptype):
    t = f"{name} {ptype}".lower()
    if any(x in t for x in ("twiser", "tweezer", "plastic form", "packet", "bean wax", "bubble")):
        return "pcs"
    return "ml"


def product_type(ptype, name):
    t = f"{ptype} {name}".lower()
    if any(x in t for x in ("color", "developer", "bleach", "chemical", "reducteur", "volume")):
        return "chemical"
    if any(x in t for x in ("twiser", "plastic", "packet", "form")):
        return "accessories"
    return "consumable"


def brand_of(name):
    n = name.strip()
    known = [
        "Spa Cylon", "Color Zone", "Natures Essence", "Natures Essence Diamond",
        "Moon Star", "Ume Care", "White rice", "Pro Tech", "Bond Fusion",
    ]
    for k in known:
        if n.lower().startswith(k.lower()):
            return k
    m = re.match(r"^([A-Za-z][A-Za-z\s&\.]+?)(?:\s+[\d\./,]|\s*$)", n)
    if m:
        b = m.group(1).strip(" -")
        if len(b) >= 2:
            return b[:120]
    parts = n.split()
    return (parts[0][:120] if parts else None)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Stock Aug 2026"]
    merged = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_name = str(row[1]).strip() if row[1] is not None else ""
        ptype = str(row[3]).strip() if row[3] is not None else ""
        if not raw_name or raw_name.lower() in ("brand", "none"):
            continue

        opening = num(row[4])
        closing = num(row[7])
        if closing is not None:
            stock = closing
        elif opening is not None:
            stock = opening
        else:
            stock = 0.0

        # Always include type in the name when present so brand-house rows stay unique
        # (e.g. "Ume Care — Foaming cleanser", "Loreal 01 — Colors").
        if ptype:
            display = f"{raw_name} — {ptype}"
        else:
            display = raw_name
        key = display.strip().lower()

        if key in merged:
            merged[key]["opening_stock"] = round(merged[key]["opening_stock"] + float(stock), 3)
            continue

        notes = []
        if ptype:
            notes.append(f"Type: {ptype}")
        notes.append("Imported from Larvendo Stock Aug 2026")

        merged[key] = {
            "name": display[:180],
            "brand": brand_of(raw_name),
            "product_type": product_type(ptype, raw_name),
            "unit": unit_for(raw_name, ptype),
            "opening_stock": round(float(stock), 3),
            "min_stock": 0,
            "status": "active",
            "notes": " | ".join(notes),
        }

    products = sorted(merged.values(), key=lambda p: p["name"].lower())
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(products, indent=2), encoding="utf-8")
    print(f"Wrote {len(products)} products → {OUT}")


if __name__ == "__main__":
    main()
